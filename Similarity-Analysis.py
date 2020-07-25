# allows you to use openpyxl functionality or code
from openpyxl import Workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

# Fill in the row index where the first row repeats
index_of_first_row = [0, 11] 

# specifically unique openpyxl // way to define Font & size of text
stft = Font(name='Consolas', size=10.5)
ft = Font(name='Consolas', size=10.5, color=colors.WHITE)

# create an empty list, variable f is where I can find the text file 
kyra_data = []
f = open(r"C:\Users\bruno\Documents\Python Scripts\Kyra PHD\data.txt", "r") # read in a notepad, in read-only mode & open the file

# loop over every line in your notepad, and add that to the list "kyra_data"
for line in f:
    kyra_data.append(line)
    #print(line)

f.close() # you telling python that you are done with the text file


print("File read and stored as list, printing list")
kyra_data = [x for x in kyra_data if len(x)>5]

#Create excel worksheet
filename = r"C:\Users\bruno\Documents\Python Scripts\Kyra PHD\output.xlsx"

workbook = Workbook()
sheet = workbook.active

column_range = []
for index, character in enumerate(kyra_data[0]):
    if index != 0 and character.isupper():
        column_range.append(index)

for i in range(0,len(kyra_data)):
    print("Row",i,"contains",len(kyra_data[i]),"characters")
    row_match_counter = [0,0,0]

    for j in range(0,len(kyra_data[i])):        
        sheet.cell(row=i+1, column=j+1).value = kyra_data[i][j]
        sheet.cell(row=i+1, column=j+1).font = stft
    
        if j>= min(column_range) and j<= max(column_range):
            # first check which row to compare against, 0 or 11
            
            if i in index_of_first_row: 
                comparator_row = i

            # then check if the row is the comparator row, if it isn't check if it a perfect or fuzzy match
            if i == comparator_row and kyra_data[i][j] != "-" and kyra_data[i][j] != " ":
                sheet.cell(row=i+1, column=j+1).fill = PatternFill(fgColor="007A5A", fill_type = "solid") # Dark color
                sheet.cell(row=i+1, column=j+1).font = ft
                row_match_counter[0] += 1

            else:
                try:
                    if ((kyra_data[comparator_row][j] == kyra_data[i][j]) and (kyra_data[i][j] != " ") and (kyra_data[i][j] != "-")):
                        #print(i,j, "equals",i-1,j, "success")
                        sheet.cell(row=i+1, column=j+1).fill = PatternFill(fgColor="007A5A", fill_type = "solid") # Dark color
                        sheet.cell(row=i+1, column=j+1).font = ft
                        row_match_counter[0] += 1

                    elif (kyra_data[i][j] in ("D","E","N","Q") and kyra_data[comparator_row][j] in ("D","E","N","Q")) or \
                    (kyra_data[i][j] in ("K","R","H") and kyra_data[comparator_row][j] in ("K","R","H")) or \
                    (kyra_data[i][j] in ("F","W","Y") and kyra_data[comparator_row][j] in ("F","W","Y")) or \
                    (kyra_data[i][j] in ("V","I","L","M") and kyra_data[comparator_row][j] in ("V","I","L","M")) or \
                    (kyra_data[i][j] in ("S","T") and kyra_data[comparator_row][j] in ("S","T")):
                        sheet.cell(row=i+1, column=j+1).fill = PatternFill(fgColor="00B082", fill_type = "solid") # Light color
                        sheet.cell(row=i+1, column=j+1).font = ft
                        row_match_counter[1] += 1

                    elif (kyra_data[i][j] != " ") and (kyra_data[i][j] != "-"):
                        row_match_counter[2] += 1
                    
                    else:
                        pass

                except():
                    pass

    sheet.cell(row=i+1, column=100).value = str(row_match_counter[0])
    sheet.cell(row=i+1, column=101).value = str(row_match_counter[1])
    sheet.cell(row=i+1, column=102).value = str(row_match_counter[2])
    sheet.cell(row=i+1, column=103).value = sum(row_match_counter)


column = 1
while column < 601:
    if column > 99:
        i = get_column_letter(column)
        sheet.column_dimensions[i].width = 3 
        column += 1
    else:
        i = get_column_letter(column)
        sheet.column_dimensions[i].width = 1.5
        column += 1

sheet.sheet_view.showGridLines = False

workbook.save(filename=filename)

print(kyra_data[0][32], type(kyra_data[0][32]))
